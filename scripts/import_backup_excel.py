"""분류 완료된 백업 엑셀(지출내역서_분류완료.xlsx)을 Supabase에 일괄 임포트한다.

원본 카드문자/카카오페이 텍스트가 없는 과거 데이터라 capture 없이 receipts에 바로
기록한다 (은행이체 수동입력과 같은 패턴). source_type='manual_other', memo에 임포트
배치 마커를 남겨 재실행해도 중복 없이 덮어쓸 수 있게 한다(먼저 같은 마커의 기존
행을 지우고 다시 넣음).

실행 전 SUPABASE_DB_URL이 설정돼 있는지 반드시 확인한다 (로컬 SQLite 테스트 DB에
실수로 153건이 들어가는 걸 방지).
"""

import re
import sys
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import connection as db_connection  # noqa: E402
from app.db.seed_categories import seed_categories  # noqa: E402
from app.services.category_service import get_category_id  # noqa: E402
from app.services.receipt_service import classify_receipt, create_receipt  # noqa: E402

EXCEL_PATH = Path(__file__).resolve().parent.parent / "지출내역서_분류완료.xlsx"
IMPORT_MARKER = "엑셀_백업_임포트_2026-08"

_FULL_DATE_RE = re.compile(r"^(\d{4})\.(\d{2})\.(\d{2})$")
_MMDD_DOW_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\s*\(.\)$")
_MMDD_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})$")
_SHIFTED_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})\s*·")


def parse_date(raw) -> str:
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = str(raw).strip()
    m = _FULL_DATE_RE.match(s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = _MMDD_DOW_RE.match(s)
    if m:
        return f"2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    m = _MMDD_RE.match(s)
    if m:
        return f"2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    raise ValueError(f"날짜 형식을 인식할 수 없습니다: {raw!r}")


def parse_time(raw) -> str | None:
    if isinstance(raw, time):
        return raw.strftime("%H:%M")
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ("", "-"):
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return None


def main() -> None:
    if not db_connection.SUPABASE_DB_URL:
        raise SystemExit("SUPABASE_DB_URL이 설정되지 않았습니다. .env를 확인하세요 (안전장치).")

    db_connection.init_db()
    seed_categories()

    conn = db_connection.get_connection()
    try:
        deleted = conn.execute("SELECT COUNT(*) AS n FROM receipts WHERE memo = ?", (IMPORT_MARKER,)).fetchone()["n"]
        conn.execute("DELETE FROM receipts WHERE memo = ?", (IMPORT_MARKER,))
        conn.commit()
    finally:
        conn.close()
    print(f"기존 임포트 배치 {deleted}건 삭제 (재실행 대비 초기화)")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]

    inserted = 0
    skipped = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        source = row[1].value  # B 거래수단
        date_raw = row[2].value  # C 날짜
        time_raw = row[3].value  # D 이용시간
        merchant = row[4].value  # E 사용처
        amount = row[5].value  # F 금액
        method = row[6].value  # G 결제방식/카드정보
        gubun = row[7].value  # H 구분
        major = row[8].value  # I 대분류
        minor = row[9].value  # J 소분류

        if merchant is None or amount is None:
            continue

        merchant = str(merchant)
        # 원본 엑셀 한 행(매듭병원 07.01)이 열이 밀려 들어와 있음: D='매듭병원', E='15:06 · 현대카드M · 일시불'
        if _SHIFTED_TIME_RE.match(merchant):
            time_raw = merchant
            merchant = str(row[3].value)

        entry_type = "expense" if gubun == "지출" else "income"
        category_id = get_category_id(entry_type, major, minor or "")
        if category_id is None:
            skipped.append((merchant, amount, gubun, major, minor, "카테고리 없음"))
            continue

        try:
            txn_date = parse_date(date_raw)
        except ValueError as e:
            skipped.append((merchant, amount, gubun, major, minor, str(e)))
            continue
        txn_time = parse_time(time_raw)
        payment_method = f"{source} / {method}" if method else source

        flow_direction = "inflow" if entry_type == "income" else "outflow"
        receipt_id = create_receipt(
            capture_id=None,
            ocr_result_id=None,
            entry_type=entry_type,
            source_type="manual_other",
            flow_direction=flow_direction,
            merchant_name=merchant,
            amount=abs(int(amount)),
            transaction_date=txn_date,
            transaction_time=txn_time,
            payment_method=payment_method,
            memo=IMPORT_MARKER,
            review_status="confirmed",
            is_manual_entry=1,
        )
        classify_receipt(receipt_id, category_id, "user", note="엑셀 백업 데이터 일괄 임포트")
        inserted += 1

    print(f"임포트 완료: {inserted}건")
    if skipped:
        print(f"건너뛴 행: {len(skipped)}건")
        for s in skipped:
            print("  ", s)


if __name__ == "__main__":
    main()
